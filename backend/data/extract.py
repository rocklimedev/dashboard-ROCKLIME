import os
import json
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from PIL import Image
from io import BytesIO
from zipfile import ZipFile

# === Config ===
EXCEL_PATH = "OnePager New MRP.xlsx"
OUTPUT_IMAGE_DIR = "img"
OUTPUT_JSON_PATH = "output.json"
IMAGE_COLUMN = 'A'  # Images are in column A

# Validate Excel file
if not os.path.exists(EXCEL_PATH):
    raise FileNotFoundError(f"Excel file not found: {EXCEL_PATH}")

# Create output directory and verify permissions
try:
    os.makedirs(OUTPUT_IMAGE_DIR, exist_ok=True)
    # Test write permission by creating a temporary file
    test_file = os.path.join(OUTPUT_IMAGE_DIR, "test.txt")
    with open(test_file, 'w') as f:
        f.write("test")
    os.remove(test_file)
except PermissionError:
    raise PermissionError(f"Cannot write to directory: {OUTPUT_IMAGE_DIR}")

def extract_images_from_xlsx(xlsx_path):
    image_map = {}
    try:
        with ZipFile(xlsx_path, 'r') as zip_ref:
            image_files = [f for f in zip_ref.namelist() if f.startswith('xl/media/')]
            for idx, img_file in enumerate(image_files):
                with zip_ref.open(img_file) as f:
                    image_map[idx + 1] = f.read()
        print(f"Extracted {len(image_map)} images from Excel")
    except Exception as e:
        print(f"⚠️ Failed to extract images from Excel: {e}")
    return image_map

def get_image_key(img):
    anchor = img.anchor
    if hasattr(anchor, 'from_'):
        row = anchor.from_.row + 1
        col = anchor.from_.col + 1
        return f"{row}_{col}"
    return None

def clean_price(price):
    """Clean price string (e.g., '₹ 4,72,950.00') to a float."""
    if isinstance(price, str):
        price = price.replace('₹', '').replace(',', '').strip()
        try:
            return float(price)
        except ValueError:
            return None
    return price

def main():
    wb = load_workbook(EXCEL_PATH, data_only=True)
    image_map = extract_images_from_xlsx(EXCEL_PATH)
    all_data = []

    for sheet in wb.worksheets:
        current_category = ""
        img_cell_map = {}

        # Step 1: Map images to cell positions
        for idx, img in enumerate(sheet._images):
            key = get_image_key(img)
            if key:
                img_cell_map[key] = idx + 1
        print(f"Image cell map for sheet {sheet.title}: {len(img_cell_map)} mappings")

        for row in sheet.iter_rows(min_row=1, values_only=False):
            row_num = row[0].row
            row_values = [
                cell.value.strip() if isinstance(cell.value, str) else cell.value
                for cell in row
            ]

            if not any(row_values):
                continue

            # Step 2: Skip header rows
            if any(
                isinstance(val, str) and val.upper() in ["NAME", "CODE", "PRICE"]
                for val in row_values
            ):
                print(f"Skipping header row {row_num}: {row_values}")
                continue

            # Step 3: Category detection
            if (
                len(row_values) >= 2
                and isinstance(row_values[1], str)
                and any(
                    keyword in row_values[1].upper()
                    for keyword in ["SHOWER TOILET", "E-BIDET", "TOILETS", "BIDET"]
                )
            ):
                current_category = row_values[1]
                print(f"Detected category in row {row_num}: {current_category}")
                continue

            # Step 4: Extract name/code/price (cols B, C, D => index 1, 2, 3)
            try:
                name = row_values[1]  # Column B
                code = row_values[2]  # Column C
                price = row_values[3]  # Column D
            except IndexError:
                print(f"Skipping row {row_num}: Insufficient columns")
                continue

            if not name or not code:
                print(f"Skipping row {row_num}: Missing name or code")
                continue

            # Clean and format price
            cleaned_price = clean_price(price)
            formatted_price = (
                f"₹ {int(cleaned_price):,}" if isinstance(cleaned_price, (int, float)) else price
            )

            # Step 5: Handle image
            col_index = column_index_from_string(IMAGE_COLUMN)
            key = f"{row_num}_{col_index}"
            image = next((i for k, i in img_cell_map.items() if k == key), None)
            image_path = None

            if image and image in image_map:
                try:
                    img_bytes = image_map[image]
                    if not img_bytes or len(img_bytes) < 100:  # Basic validation
                        print(f"⚠️ Invalid image data for code {code} (row: {row_num})")
                    else:
                        img = Image.open(BytesIO(img_bytes))
                        ext = img.format.lower() if img.format else "png"
                        filename = f"{code}.{ext}"
                        filepath = os.path.join(OUTPUT_IMAGE_DIR, filename)
                        img.save(filepath, quality=95)  # Ensure high quality
                        if os.path.exists(filepath):
                            image_path = f"./img/{filename}"
                            print(f"Saved image for code {code}: {image_path}")
                        else:
                            print(f"⚠️ Failed to verify saved image for code {code}")
                except (Image.UnidentifiedImageError, OSError) as e:
                    print(f"⚠️ Invalid image format for code {code}: {e}")
                except PermissionError as e:
                    print(f"⚠️ Permission error saving image for code {code}: {e}")
                except Exception as e:
                    print(f"⚠️ Failed to save image for code {code}: {e}")
            else:
                print(f"⚠️ Image not found for code {code} (row: {row_num})")
                image_path = "./img/placeholder.png"

            # Step 6: Push data to final array
            all_data.append({
                "sheet": sheet.title,
                "category": current_category,
                "name": name,
                "code": code,
                "price": formatted_price,
                "image_path": image_path
            })

    # Write to JSON
    try:
        with open(OUTPUT_JSON_PATH, 'w', encoding='utf-8') as f:
            json.dump(all_data, f, indent=2, ensure_ascii=False)
        print(f"✅ Done. Extracted {len(all_data)} items to {OUTPUT_JSON_PATH}")
    except Exception as e:
        print(f"⚠️ Failed to write JSON file: {e}")

if __name__ == "__main__":
    main()